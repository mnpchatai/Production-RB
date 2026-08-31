#!/usr/bin/env python3
"""
แปลงไฟล์ "ระบบวางแผนการสั่งงานแผนก RB V1.7.1" (.xlsm) เป็น CSV สำหรับนำเข้า Supabase

ใช้งาน:
    pip install openpyxl
    python3 etl.py --out data \
        --book TOY --name "สมุดงาน TOY (6926)" --file "3.ระบบ...TOY (6926).xlsm" \
        --book IND --name "สมุดงานยางอุตสาหกรรม (26)" --file "3.ระบบ...(26).xlsm"

ได้ไฟล์ CSV ชุดเดียวกับตารางใน schema.sql — นำเข้าตามลำดับใน README.md

หมายเหตุการอ่านชีต (ยึดตามไฟล์ V1.7.1):
    Master_Rubber      หัวตารางแถว 3   ข้อมูลเริ่มแถว 4    คอลัมน์ A:F
    DATA STANDARD      หัวตารางแถว 4   ข้อมูลเริ่มแถว 5    คอลัมน์ A:M
    DATA BOM           หัวตารางแถว 2   ข้อมูลเริ่มแถว 3    คอลัมน์ A:T
    Chronicle_Working  หัวตารางแถว 2   ข้อมูลเริ่มแถว 3    คอลัมน์ A:R + S:AX เก็บเป็น jsonb
    System             บล็อกลอย: A:B วันที่ / D:G สี / I:J ลูกค้า / L:M รหัสสินค้า
    LogSheet           หัวตารางแถว 3   ข้อมูลเริ่มแถว 4
คอลัมน์ N:AA ของ DATA STANDARD (มาตรฐานฉีด/อบ/HCM) ว่างทั้งสองเล่ม จึงไม่ถูกดึง
"""

import argparse, csv, datetime, json, os, re, sys
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("ต้องติดตั้งก่อน: pip install openpyxl")

# ค่าที่ถือว่าไม่มีข้อมูล — Excel เก็บ error string กับขีดคั่นปนมาด้วย
ERRORS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!"}


def txt(v):
    """ค่าดิบ -> str ที่สะอาด หรือ None ถ้าว่าง/เป็น error/เป็นขีดล้วน"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    # เลขใบสั่งผลิตบางแถวติดเครื่องหมาย ' นำหน้ามาจากการพิมพ์ใน Excel
    s = s.lstrip("'").strip()
    if not s or s in ERRORS:
        return None
    if set(s) <= set("-–— "):          # '-', '-----', '- - -'
        return None
    return s


def num(v):
    """ค่าดิบ -> float หรือ None (ตัดคอมมา, ตัด % ออกโดยไม่หาร)"""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = txt(v)
    if s is None:
        return None
    s = s.replace(",", "").rstrip("%").strip()
    try:
        return float(s)
    except ValueError:
        return None


def date(v):
    """ค่าดิบ -> 'YYYY-MM-DD' หรือ None"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    s = txt(v)
    if s and re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    return None


def cell(row, i):
    return row[i] if i < len(row) else None


def rows_from(ws, skip):
    """คืนคู่ (เลขแถวจริงในชีต, tuple ค่า) โดยข้ามหัวตาราง skip แถว"""
    for n, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if n <= skip:
            continue
        yield n, row


# --------------------------------------------------------------------------- ชีต

def read_formulas(ws, book):
    """Master_Rubber -> rb_formulas"""
    out, seen = [], set()
    for _, r in rows_from(ws, 3):
        code = txt(cell(r, 1))
        if not code or code in seen:
            continue
        seen.add(code)
        out.append({
            "book_code": book, "seq": int(num(cell(r, 0)) or 0) or None,
            "formula_code": code,
            "batch_weight_kg": num(cell(r, 2)),
            "strands_per_batch": num(cell(r, 3)),
            "weight_per_strand_kg": num(cell(r, 4)),
            "min_add_weight_kg": num(cell(r, 5)),
        })
    return out


def read_item_standards(ws, book):
    """DATA STANDARD -> rb_item_standards (ยุบรหัสซ้ำเหลือรหัสละแถว)"""
    groups = defaultdict(list)
    for _, r in rows_from(ws, 4):
        code = txt(cell(r, 6))
        if not code:
            continue
        groups[code].append(r)

    out = []
    for code, rs in groups.items():
        # ชีตต้นทางมีรหัสเดิมซ้ำได้หลายร้อยแถว ต่างกันแค่ชื่อที่พ่วงสี
        # เลือกชุดค่าที่พบบ่อยที่สุดเป็นตัวแทน ชื่อก็เลือกที่พบบ่อยที่สุดเช่นกัน
        core = Counter(
            tuple(txt(cell(r, i)) for i in (1, 2, 3, 4, 5, 8, 9, 10, 11, 12)) for r in rs
        ).most_common(1)[0][0]
        names = Counter(txt(cell(r, 7)) for r in rs if txt(cell(r, 7)))
        seqs = [num(cell(r, 0)) for r in rs if num(cell(r, 0)) is not None]
        out.append({
            "book_code": book,
            "seq": int(min(seqs)) if seqs else None,
            "dept": core[0], "rubber_type": core[1],
            "length_txt": core[2], "hole_txt": core[3], "outer_txt": core[4],
            "item_code": code,
            "item_name": names.most_common(1)[0][0] if names else None,
            "weight_per_strand_g": num(core[5]),
            "head_allowance_kg": num(core[6]),
            "joint_scrap_pct": num(core[7]),
            "general_scrap_pct": num(core[8]),
            "formula_code": core[9],
            "name_variants": len(names),
            "source_rows": len(rs),
        })
    out.sort(key=lambda d: (d["seq"] is None, d["seq"] or 0))
    return out


def read_bom(ws, book):
    """DATA BOM -> rb_bom_lines (คอลัมน์ A:T เท่านั้น, V:AP กับ AR:BN เป็นบล็อกคำนวณ)"""
    out = []
    for n, r in rows_from(ws, 2):
        item, code = txt(cell(r, 0)), txt(cell(r, 2))
        if not item or not code:
            continue
        out.append({
            "book_code": book, "src_row": n,
            "product_item": item,
            "rb_weight_g": num(cell(r, 1)),
            "rb_code": code,
            "dept_code": txt(cell(r, 3)),
            "formula_code": txt(cell(r, 4)),
            "rubber_type": txt(cell(r, 5)),
            "color_code": txt(cell(r, 6)),
            "length_txt": txt(cell(r, 7)),
            "hole_txt": txt(cell(r, 8)),
            "outer_txt": txt(cell(r, 9)),
            "rb_name": txt(cell(r, 10)),
            "producer_dept": txt(cell(r, 11)),
            "qty_per_set": num(cell(r, 12)),
            "cut_length": num(cell(r, 13)),
            "cut_length_unit": txt(cell(r, 14)),
            "pcs_per_strand": num(cell(r, 15)),
            "pcs_unit": txt(cell(r, 16)),
            "qty_per_gr": num(cell(r, 17)),
            "qty_per_gr_unit": txt(cell(r, 18)),
            "rb_uom": txt(cell(r, 19)),
        })
    return out


# คอลัมน์ A:R ของ Chronicle_Working -> ชื่อฟิลด์ในตาราง
WO_COLS = [
    (0, "seq", num), (1, "customer_name", txt), (2, "production_order_no", txt),
    (3, "fg_code", txt), (4, "rb_code", txt), (5, "rb_name", txt), (6, "color", txt),
    (7, "qty_kg", num), (8, "mixing_strands", num), (9, "rb_strands", num),
    (10, "head_allowance_kg", num), (11, "joint_allowance_kg", num),
    (12, "general_scrap_kg", num), (13, "other_dept_scrap_kg", num),
    (14, "mixing_start", date), (15, "injection_due", date), (16, "hcm_due", date),
    (17, "work_order_no", txt),
]


def read_work_orders(ws, book):
    """Chronicle_Working -> rb_work_order_lines (S:AX ยัดลง std_snapshot)"""
    header = None
    out = []
    for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
        if n == 2:
            header = r
            continue
        if n < 3 or header is None:
            continue
        rec = {"book_code": book, "src_row": n}
        for i, field, conv in WO_COLS:
            rec[field] = conv(cell(r, i))
        # ไม่มีทั้งเลขใบสั่งผลิต ลูกค้า และรหัส RB = แถวเปล่าที่มีแค่เลขลำดับ
        if not (rec["production_order_no"] or rec["customer_name"] or rec["rb_code"]):
            continue
        rec["seq"] = int(rec["seq"]) if rec["seq"] is not None else None
        snap = {}
        for i in range(18, len(r)):
            v = txt(cell(r, i))
            if v is None:
                continue
            key = txt(cell(header, i)) or f"col{i + 1}"
            snap[key] = v
        rec["std_snapshot"] = json.dumps(snap, ensure_ascii=False)
        out.append(rec)
    return out


def read_system(ws, book):
    """System -> (colors, customers, products, settings)

    ชีตนี้วางเป็นบล็อกลอยข้าง ๆ กัน แถว 4-9 คือค่าที่ค้างอยู่บนฟอร์ม
    แถว 10 เป็นหัวตารางของทุกบล็อก ข้อมูลเริ่มแถว 11
    """
    rows = list(ws.iter_rows(values_only=True))

    def g(r, i):
        return cell(rows[r], i) if r < len(rows) else None

    colors, customers, products = [], [], []
    seen_c, seen_cu, seen_p = set(), set(), set()
    for r in range(10, len(rows)):
        code = txt(g(r, 4))
        if code and code not in seen_c:
            seen_c.add(code)
            colors.append({"book_code": book, "seq": int(num(g(r, 3)) or 0) or None,
                           "color_code": code, "color_name": txt(g(r, 5)),
                           "color_group": txt(g(r, 6))})
        cu = txt(g(r, 9))
        if cu and cu not in seen_cu:
            seen_cu.add(cu)
            customers.append({"book_code": book, "seq": int(num(g(r, 8)) or 0) or None,
                              "customer_name": cu})
        p = txt(g(r, 12))
        if p and p not in seen_p:
            seen_p.add(p)
            products.append({"book_code": book, "seq": int(num(g(r, 11)) or 0) or None,
                             "item_code": p})

    settings = {
        "กำหนดผลิตเสร็จตียาง": date(g(4, 1)),
        "กำหนดผลิตเสร็จฉีดยาง": date(g(6, 1)),
        "กำหนดผลิตเสร็จ HCM": date(g(8, 1)),
        "สีที่ 1": txt(g(6, 4)),
        "สีที่ 2": txt(g(8, 4)),
        "ลูกค้าที่เลือกอยู่": txt(g(8, 9)),
        "รหัสสินค้าที่เลือกอยู่": txt(g(8, 12)),
    }
    return colors, customers, products, {k: v for k, v in settings.items() if v}


def read_doc_log(ws, book):
    """LogSheet -> (แถวทะเบียนเอกสาร, เลขที่เอกสารที่หัวชีต)"""
    rows = list(ws.iter_rows(values_only=True))
    doc_no = txt(cell(rows[1], 8)) if len(rows) > 1 else None
    out = []
    for n, r in rows_from(ws, 3):
        seq = num(cell(r, 0))
        if seq is None:
            continue
        rec = {"book_code": book, "seq": int(seq),
               "sent_date": date(cell(r, 1)), "work_order_no": txt(cell(r, 2)),
               "job_type": txt(cell(r, 3)), "doc_count": txt(cell(r, 4)),
               "customer_name": txt(cell(r, 5)), "pi_order_no": txt(cell(r, 6)),
               "sender": txt(cell(r, 7)), "receiver": txt(cell(r, 8))}
        if any(rec[k] for k in rec if k not in ("book_code", "seq")):
            out.append(rec)
    return out, doc_no


# --------------------------------------------------------------------------- main

SHEET = {"formulas": "Master_Rubber", "standards": "DATA STANDARD", "bom": "DATA BOM",
         "wo": "Chronicle_Working", "system": "System", "log": "LogSheet "}


def sheet(wb, key):
    """หาชีตแบบไม่สนช่องว่างท้ายชื่อ (ชีต 'LogSheet ' มีช่องว่างจริง)"""
    want = SHEET[key].strip()
    for name in wb.sheetnames:
        if name.strip() == want:
            return wb[name]
    raise KeyError(f"ไม่พบชีต {SHEET[key]!r} — มีชีต: {wb.sheetnames}")


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in fields})


TABLES = {
    "rb_books": ["code", "name", "source_file", "app_version", "doc_no", "settings"],
    "rb_colors": ["book_code", "seq", "color_code", "color_name", "color_group"],
    "rb_customers": ["book_code", "seq", "customer_name"],
    "rb_products": ["book_code", "seq", "item_code"],
    "rb_formulas": ["book_code", "seq", "formula_code", "batch_weight_kg",
                    "strands_per_batch", "weight_per_strand_kg", "min_add_weight_kg"],
    "rb_item_standards": ["book_code", "seq", "dept", "rubber_type", "length_txt",
                          "hole_txt", "outer_txt", "item_code", "item_name",
                          "weight_per_strand_g", "head_allowance_kg", "joint_scrap_pct",
                          "general_scrap_pct", "formula_code", "name_variants",
                          "source_rows"],
    "rb_bom_lines": ["book_code", "src_row", "product_item", "rb_weight_g", "rb_code",
                     "dept_code", "formula_code", "rubber_type", "color_code",
                     "length_txt", "hole_txt", "outer_txt", "rb_name", "producer_dept",
                     "qty_per_set", "cut_length", "cut_length_unit", "pcs_per_strand",
                     "pcs_unit", "qty_per_gr", "qty_per_gr_unit", "rb_uom"],
    "rb_work_order_lines": ["book_code", "src_row", "seq", "customer_name",
                            "production_order_no", "fg_code", "rb_code", "rb_name",
                            "color", "qty_kg", "mixing_strands", "rb_strands",
                            "head_allowance_kg", "joint_allowance_kg",
                            "general_scrap_kg", "other_dept_scrap_kg", "mixing_start",
                            "injection_due", "hcm_due", "work_order_no",
                            "std_snapshot"],
    "rb_doc_log": ["book_code", "seq", "sent_date", "work_order_no", "job_type",
                   "doc_count", "customer_name", "pi_order_no", "sender", "receiver"],
}


# โหมดในเครื่องของแอปอ่านแค่แปดตารางนี้ และใช้ไม่ครบทุกคอลัมน์
# ตัดคอลัมน์ที่ไม่ได้ใช้กับค่า null ออก ช่วยให้ไฟล์เล็กลงจาก 7.2 MB เหลือ ~5 MB
JSON_TABLES = {
    "rb_books":            ["code", "name", "source_file", "app_version", "doc_no"],
    "rb_colors":           ["book_code", "seq", "color_code", "color_name", "color_group"],
    "rb_customers":        ["book_code", "seq", "customer_name"],
    "rb_products":         ["book_code", "seq", "item_code"],
    "rb_formulas":         ["book_code", "seq", "formula_code", "batch_weight_kg",
                            "strands_per_batch", "weight_per_strand_kg", "min_add_weight_kg"],
    "rb_item_standards":   ["book_code", "item_code", "item_name", "dept", "rubber_type",
                            "length_txt", "hole_txt", "outer_txt", "weight_per_strand_g",
                            "head_allowance_kg", "joint_scrap_pct", "general_scrap_pct",
                            "formula_code"],
    "rb_bom_lines":        ["book_code", "src_row", "product_item", "rb_code", "rb_code_std",
                            "rb_name", "color_code", "formula_code", "rb_weight_g", "qty_per_set",
                            "cut_length", "cut_length_unit", "pcs_per_strand", "pcs_unit",
                            "qty_per_gr", "rb_uom"],
    "rb_work_order_lines": ["id", "book_code", "src_row", "seq", "production_order_no",
                            "work_order_no", "customer_name", "fg_code", "rb_code", "rb_name",
                            "color", "qty_kg", "mixing_strands", "rb_strands",
                            "head_allowance_kg", "joint_allowance_kg", "general_scrap_kg",
                            "mixing_start", "injection_due", "hcm_due"],
}


def write_json(path, acc):
    """ไฟล์เดียวสำหรับโหมดในเครื่องของแอป (ตั้งค่าระบบ → ข้อมูล → นำเข้า)"""
    out = {}
    for table, fields in JSON_TABLES.items():
        rows = []
        for i, r in enumerate(acc[table], start=1):
            rec = {k: r[k] for k in fields if r.get(k) not in (None, "")}
            if table == "rb_bom_lines":
                # ในฐานข้อมูลคอลัมน์นี้เป็น generated column — ไฟล์ JSON ต้องประกอบเอง
                # ให้ตรงรูปแบบเดียวกัน ไม่งั้น BOM จะ join กับมาตรฐานชิ้นงานไม่ติด
                rec["rb_code_std"] = "-".join(
                    str(r.get(k) or "") for k in
                    ("dept_code", "formula_code", "rubber_type", "length_txt", "hole_txt", "outer_txt"))
            # ฐานข้อมูลออก id ให้เอง แต่ไฟล์ JSON ไม่มีใครออกให้ ต้องใส่มาเอง
            if "id" in fields and "id" not in rec:
                rec["id"] = i
            rows.append(rec)
        out[table] = rows
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, separators=(",", ":"))
    return sum(len(v) for v in out.values())


def main():
    ap = argparse.ArgumentParser(description="แปลง .xlsm ของระบบวางแผน RB เป็น CSV สำหรับ Supabase")
    ap.add_argument("--out", default="data", help="โฟลเดอร์ปลายทางของไฟล์ CSV")
    ap.add_argument("--book", action="append", required=True, metavar="CODE",
                    help="รหัสสมุดงาน เช่น TOY (ใส่ซ้ำได้ตามจำนวนไฟล์)")
    ap.add_argument("--name", action="append", default=[], metavar="NAME",
                    help="ชื่อสมุดงานที่จะแสดง (เรียงคู่กับ --book)")
    ap.add_argument("--file", action="append", required=True, metavar="XLSM",
                    help="พาธไฟล์ .xlsm (เรียงคู่กับ --book)")
    ap.add_argument("--json", metavar="PATH",
                    help="เขียนไฟล์ JSON ไฟล์เดียวสำหรับโหมดในเครื่องของแอปด้วย")
    a = ap.parse_args()

    if len(a.book) != len(a.file):
        sys.exit("จำนวน --book กับ --file ต้องเท่ากัน")
    names = a.name + a.book[len(a.name):]

    os.makedirs(a.out, exist_ok=True)
    acc = {t: [] for t in TABLES}

    for code, name, path in zip(a.book, names, a.file):
        print(f"\n=== {code}: {os.path.basename(path)}")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        colors, customers, products, settings = read_system(sheet(wb, "system"), code)
        log, doc_no = read_doc_log(sheet(wb, "log"), code)

        acc["rb_books"].append({
            "code": code, "name": name, "source_file": os.path.basename(path),
            "app_version": "V1.7.1", "doc_no": doc_no,
            "settings": json.dumps(settings, ensure_ascii=False),
        })
        acc["rb_colors"] += colors
        acc["rb_customers"] += customers
        acc["rb_products"] += products
        acc["rb_doc_log"] += log
        acc["rb_formulas"] += read_formulas(sheet(wb, "formulas"), code)
        acc["rb_item_standards"] += read_item_standards(sheet(wb, "standards"), code)
        acc["rb_bom_lines"] += read_bom(sheet(wb, "bom"), code)
        acc["rb_work_order_lines"] += read_work_orders(sheet(wb, "wo"), code)
        wb.close()

        for t in TABLES:
            n = sum(1 for r in acc[t] if r.get("book_code", r.get("code")) == code)
            print(f"    {t:22s} {n:6d}")

    print()
    for t, fields in TABLES.items():
        p = os.path.join(a.out, f"{t}.csv")
        write_csv(p, acc[t], fields)
        print(f"เขียน {p}  ({len(acc[t])} แถว, {os.path.getsize(p) / 1024:.0f} KB)")

    if a.json:
        n = write_json(a.json, acc)
        print(f"\nเขียน {a.json}  ({n} แถว, {os.path.getsize(a.json) / 1024 / 1024:.2f} MB)"
              f" — นำเข้าที่แอป: ตั้งค่าระบบ → ข้อมูล → นำเข้า rb_data.json")


if __name__ == "__main__":
    main()
